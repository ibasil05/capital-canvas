"""
Excel export functionality for financial models.
"""

import io
import os
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, Any, List, Optional
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.chart import LineChart, Reference, BarChart, Series

from backend.config import config

class ExcelExport:
    """Excel export handler for financial models"""
    
    # Default template path
    TEMPLATE_DIR = Path(__file__).parent.parent.parent / "templates"
    
    def __init__(self, model_data: Dict[str, Any], ticker: str, company_name: str):
        """
        Initialize Excel export handler.
        
        Args:
            model_data: Financial model data
            ticker: Company ticker
            company_name: Company name
        """
        self.model_data = model_data
        self.ticker = ticker
        self.company_name = company_name
        
        # Create workbook
        self.workbook = openpyxl.Workbook()
        
        # Define styles
        self._define_styles()
    
    def _define_styles(self):
        """Define Excel styles for consistent formatting"""
        # Fonts
        self.title_font = Font(name='Arial', size=14, bold=True)
        self.header_font = Font(name='Arial', size=12, bold=True)
        self.subheader_font = Font(name='Arial', size=11, bold=True, italic=True)
        self.normal_font = Font(name='Arial', size=10)
        
        # Fills
        self.header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        self.highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # Borders
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alignments
        self.center_align = Alignment(horizontal='center')
        self.right_align = Alignment(horizontal='right')
    
    def generate(self) -> bytes:
        """
        Generate Excel file containing the financial model.
        
        Returns:
            Excel file as bytes
        """
        # Remove default worksheet
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)
        
        # Create sheets
        self._create_summary_sheet()
        self._create_assumptions_sheet()
        self._create_income_statement_sheet()
        self._create_balance_sheet_sheet()
        self._create_cash_flow_sheet()
        self._create_valuation_sheet()
        self._create_capital_structure_sheet()
        
        # Save to bytes IO
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _create_summary_sheet(self):
        """Create the summary dashboard sheet"""
        ws = self.workbook.create_sheet("Summary")
        
        # Set column widths
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Title and company info
        ws['A1'] = f"{self.company_name} ({self.ticker}) - Financial Model"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:I1')
        
        # Valuation summary section
        ws['A3'] = "Valuation Summary"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill
        ws.merge_cells('A3:I3')
        
        # DCF valuation
        ws['A4'] = "DCF Valuation"
        ws['A4'].font = self.subheader_font
        
        ws['A5'] = "Enterprise Value"
        ws['B5'] = self.model_data.get("dcf_valuation", {}).get("enterprise_value", 0)
        ws['B5'].number_format = '"$"#,##0.0,,"B"'
        
        ws['A6'] = "Equity Value"
        ws['B6'] = self.model_data.get("dcf_valuation", {}).get("equity_value", 0)
        ws['B6'].number_format = '"$"#,##0.0,,"B"'
        
        ws['A7'] = "Share Price"
        ws['B7'] = self.model_data.get("dcf_valuation", {}).get("price_per_share", 0)
        ws['B7'].number_format = '"$"#,##0.00'
        
        # Trading comps valuation
        ws['C4'] = "Trading Comps"
        ws['C4'].font = self.subheader_font
        
        ws['C5'] = "Enterprise Value"
        ws['D5'] = self.model_data.get("trading_comps_valuation", {}).get("enterprise_value", 0)
        ws['D5'].number_format = '"$"#,##0.0,,"B"'
        
        ws['C6'] = "Equity Value"
        ws['D6'] = self.model_data.get("trading_comps_valuation", {}).get("equity_value", 0)
        ws['D6'].number_format = '"$"#,##0.0,,"B"'
        
        ws['C7'] = "Share Price"
        ws['D7'] = self.model_data.get("trading_comps_valuation", {}).get("price_per_share", 0)
        ws['D7'].number_format = '"$"#,##0.00'
        
        # LBO valuation
        ws['E4'] = "LBO Analysis"
        ws['E4'].font = self.subheader_font
        
        ws['E5'] = "Entry EV"
        ws['F5'] = self.model_data.get("lbo_valuation", {}).get("entry_enterprise_value", 0)
        ws['F5'].number_format = '"$"#,##0.0,,"B"'
        
        ws['E6'] = "Exit EV"
        ws['F6'] = self.model_data.get("lbo_valuation", {}).get("exit_enterprise_value", 0)
        ws['F6'].number_format = '"$"#,##0.0,,"B"'
        
        ws['E7'] = "Equity IRR"
        ws['F7'] = self.model_data.get("lbo_valuation", {}).get("equity_irr", 0)
        ws['F7'].number_format = '0.0%'
        
        # Add some charts for visual representation
        self._add_summary_charts(ws)
    
    def _add_summary_charts(self, ws):
        """Add charts to the summary sheet"""
        # Growth chart
        income_data = self.model_data.get("income_statement", {})
        years = list(range(5))  # Assume 5 years of forecasts
        
        if isinstance(income_data, dict) and "revenue" in income_data:
            revenue = [income_data["revenue"].get(str(year), 0) for year in years]
            
            # Create revenue growth chart
            chart = LineChart()
            chart.title = "Revenue Growth"
            chart.style = 10
            chart.y_axis.title = "Revenue ($B)"
            chart.x_axis.title = "Year"
            
            # Add data
            data = Reference(ws, min_col=8, min_row=15, max_row=15+len(years), max_col=8)
            chart.add_data(data, titles_from_data=True)
            
            # Add revenue data for chart
            ws['H14'] = "Revenue Chart Data"
            ws['H14'].font = self.header_font
            ws['H15'] = "Revenue ($B)"
            
            for i, year in enumerate(years):
                ws[f'G{16+i}'] = f"Year {year+1}"
                ws[f'H{16+i}'] = revenue[i] / 1_000_000_000  # Convert to billions
            
            # Place chart
            ws.add_chart(chart, "A10")
    
    def _create_assumptions_sheet(self):
        """Create the assumptions sheet"""
        ws = self.workbook.create_sheet("Assumptions")
        
        # Set column widths
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Title
        ws['A1'] = "Model Assumptions"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:I1')
        
        # Growth assumptions
        ws['A3'] = "Growth Assumptions"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill
        ws.merge_cells('A3:C3')
        
        growth_rates = self.model_data.get("growth_assumptions", {}).get("revenue_growth_rates", [0.05, 0.04, 0.03, 0.03, 0.02])
        
        ws['A4'] = "Revenue Growth Rates"
        for i, rate in enumerate(growth_rates):
            ws[f'{chr(66+i)}4'] = f"Year {i+1}"
            ws[f'{chr(66+i)}4'].font = self.subheader_font
            ws[f'{chr(66+i)}5'] = rate
            ws[f'{chr(66+i)}5'].number_format = '0.0%'
        
        # Margin assumptions
        ws['A7'] = "Margin Assumptions"
        ws['A7'].font = self.header_font
        ws['A7'].fill = self.header_fill
        ws.merge_cells('A7:C7')
        
        gross_margins = self.model_data.get("margin_assumptions", {}).get("gross_margins", [0.6, 0.6, 0.61, 0.61, 0.62])
        ebitda_margins = self.model_data.get("margin_assumptions", {}).get("ebitda_margins", [0.25, 0.25, 0.26, 0.26, 0.27])
        
        ws['A8'] = "Gross Margins"
        for i, margin in enumerate(gross_margins):
            ws[f'{chr(66+i)}8'] = f"Year {i+1}"
            ws[f'{chr(66+i)}8'].font = self.subheader_font
            ws[f'{chr(66+i)}9'] = margin
            ws[f'{chr(66+i)}9'].number_format = '0.0%'
        
        ws['A10'] = "EBITDA Margins"
        for i, margin in enumerate(ebitda_margins):
            ws[f'{chr(66+i)}10'] = f"Year {i+1}"
            ws[f'{chr(66+i)}10'].font = self.subheader_font
            ws[f'{chr(66+i)}11'] = margin
            ws[f'{chr(66+i)}11'].number_format = '0.0%'
        
        # Working capital assumptions
        ws['A13'] = "Working Capital Assumptions"
        ws['A13'].font = self.header_font
        ws['A13'].fill = self.header_fill
        ws.merge_cells('A13:C13')
        
        ws['A14'] = "Receivable Days"
        ws['B14'] = self.model_data.get("working_capital_assumptions", {}).get("receivable_days", 45)
        
        ws['A15'] = "Inventory Days"
        ws['B15'] = self.model_data.get("working_capital_assumptions", {}).get("inventory_days", 60)
        
        ws['A16'] = "Payable Days"
        ws['B16'] = self.model_data.get("working_capital_assumptions", {}).get("payable_days", 30)
        
        # Valuation assumptions
        ws['A18'] = "Valuation Assumptions"
        ws['A18'].font = self.header_font
        ws['A18'].fill = self.header_fill
        ws.merge_cells('A18:C18')
        
        ws['A19'] = "WACC"
        ws['B19'] = self.model_data.get("valuation_assumptions", {}).get("discount_rate", 0.1)
        ws['B19'].number_format = '0.0%'
        
        ws['A20'] = "Terminal Growth Rate"
        ws['B20'] = self.model_data.get("valuation_assumptions", {}).get("terminal_growth_rate", 0.02)
        ws['B20'].number_format = '0.0%'
        
        ws['A21'] = "EV/EBITDA Multiple"
        ws['B21'] = self.model_data.get("valuation_assumptions", {}).get("ev_to_ebitda_multiple", 8.0)
        ws['B21'].number_format = '0.0x'
        
        ws['A22'] = "Tax Rate"
        ws['B22'] = self.model_data.get("valuation_assumptions", {}).get("tax_rate", 0.21)
        ws['B22'].number_format = '0.0%'
    
    def _create_income_statement_sheet(self):
        """Create the income statement sheet"""
        ws = self.workbook.create_sheet("Income Statement")
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Title
        ws['A1'] = f"{self.company_name} ({self.ticker}) - Income Statement"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:G1')
        
        # Headers
        ws['A3'] = "In millions, USD"
        ws['A3'].font = self.subheader_font
        
        # Year headers
        for i in range(6):  # Historical + 5 years
            col = chr(66 + i)
            if i == 0:
                ws[f'{col}3'] = "Historical"
            else:
                ws[f'{col}3'] = f"Year {i}"
            ws[f'{col}3'].font = self.header_font
            ws[f'{col}3'].fill = self.header_fill
            ws[f'{col}3'].alignment = self.center_align
        
        # Income statement line items
        income_data = self.model_data.get("income_statement", {})
        
        row = 4
        line_items = [
            ("Revenue", "revenue"),
            ("Cost of Revenue", "cost_of_revenue"),
            ("Gross Profit", "gross_profit"),
            ("Gross Margin", "gross_margin"),
            ("Operating Expenses", "operating_expenses"),
            ("EBITDA", "ebitda"),
            ("EBITDA Margin", "ebitda_margin"),
            ("Depreciation & Amortization", "depreciation"),
            ("Operating Income (EBIT)", "operating_income"),
            ("Interest Expense", "interest_expense"),
            ("Income Before Taxes", "income_before_tax"),
            ("Income Taxes", "taxes"),
            ("Net Income", "net_income"),
            ("Net Margin", "net_margin")
        ]
        
        for label, key in line_items:
            ws[f'A{row}'] = label
            
            # Add data for each year
            if isinstance(income_data, dict) and key in income_data:
                for i in range(6):  # Historical + 5 years
                    col = chr(66 + i)
                    value = income_data[key].get(str(i), 0)
                    ws[f'{col}{row}'] = value
                    
                    # Format percentages
                    if "margin" in key:
                        ws[f'{col}{row}'].number_format = '0.0%'
                    else:
                        ws[f'{col}{row}'].number_format = '"$"#,##0.0,,"M"'
            
            # Apply styles
            if key in ["revenue", "gross_profit", "ebitda", "operating_income", "net_income"]:
                ws[f'A{row}'].font = self.subheader_font
                for i in range(6):
                    col = chr(66 + i)
                    ws[f'{col}{row}'].font = self.subheader_font
            
            row += 1
    
    def _create_balance_sheet_sheet(self):
        """Create the balance sheet sheet"""
        ws = self.workbook.create_sheet("Balance Sheet")
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Title
        ws['A1'] = f"{self.company_name} ({self.ticker}) - Balance Sheet"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:G1')
        
        # Headers
        ws['A3'] = "In millions, USD"
        ws['A3'].font = self.subheader_font
        
        # Year headers
        for i in range(6):  # Historical + 5 years
            col = chr(66 + i)
            if i == 0:
                ws[f'{col}3'] = "Historical"
            else:
                ws[f'{col}3'] = f"Year {i}"
            ws[f'{col}3'].font = self.header_font
            ws[f'{col}3'].fill = self.header_fill
            ws[f'{col}3'].alignment = self.center_align
        
        # Balance sheet line items
        balance_data = self.model_data.get("balance_sheet", {})
        
        # Assets
        row = 4
        ws[f'A{row}'] = "Assets"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        asset_items = [
            ("Cash and Equivalents", "cash"),
            ("Accounts Receivable", "accounts_receivable"),
            ("Inventory", "inventory"),
            ("Other Current Assets", "other_current_assets"),
            ("Total Current Assets", "total_current_assets"),
            ("Property, Plant & Equipment", "fixed_assets"),
            ("Intangible Assets", "intangible_assets"),
            ("Other Non-Current Assets", "other_non_current_assets"),
            ("Total Non-Current Assets", "total_non_current_assets"),
            ("Total Assets", "total_assets")
        ]
        
        for label, key in asset_items:
            ws[f'A{row}'] = label
            
            # Add data for each year
            if isinstance(balance_data, dict) and key in balance_data:
                for i in range(6):  # Historical + 5 years
                    col = chr(66 + i)
                    value = balance_data[key].get(str(i), 0)
                    ws[f'{col}{row}'] = value
                    ws[f'{col}{row}'].number_format = '"$"#,##0.0,,"M"'
            
            # Apply styles
            if key in ["total_current_assets", "total_non_current_assets", "total_assets"]:
                ws[f'A{row}'].font = self.subheader_font
                for i in range(6):
                    col = chr(66 + i)
                    ws[f'{col}{row}'].font = self.subheader_font
            
            row += 1
        
        # Liabilities and Equity
        row += 1
        ws[f'A{row}'] = "Liabilities and Equity"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        liability_items = [
            ("Accounts Payable", "accounts_payable"),
            ("Short-Term Debt", "short_term_debt"),
            ("Other Current Liabilities", "other_current_liabilities"),
            ("Total Current Liabilities", "total_current_liabilities"),
            ("Long-Term Debt", "long_term_debt"),
            ("Other Non-Current Liabilities", "other_non_current_liabilities"),
            ("Total Non-Current Liabilities", "total_non_current_liabilities"),
            ("Total Liabilities", "total_liabilities"),
            ("Common Stock", "common_stock"),
            ("Retained Earnings", "retained_earnings"),
            ("Total Equity", "total_equity"),
            ("Total Liabilities and Equity", "total_liabilities_and_equity")
        ]
        
        for label, key in liability_items:
            ws[f'A{row}'] = label
            
            # Add data for each year
            if isinstance(balance_data, dict) and key in balance_data:
                for i in range(6):  # Historical + 5 years
                    col = chr(66 + i)
                    value = balance_data[key].get(str(i), 0)
                    ws[f'{col}{row}'] = value
                    ws[f'{col}{row}'].number_format = '"$"#,##0.0,,"M"'
            
            # Apply styles
            if key in ["total_current_liabilities", "total_non_current_liabilities", "total_liabilities", "total_equity", "total_liabilities_and_equity"]:
                ws[f'A{row}'].font = self.subheader_font
                for i in range(6):
                    col = chr(66 + i)
                    ws[f'{col}{row}'].font = self.subheader_font
            
            row += 1
    
    def _create_cash_flow_sheet(self):
        """Create the cash flow statement sheet"""
        ws = self.workbook.create_sheet("Cash Flow")
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Title
        ws['A1'] = f"{self.company_name} ({self.ticker}) - Cash Flow Statement"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:G1')
        
        # Headers
        ws['A3'] = "In millions, USD"
        ws['A3'].font = self.subheader_font
        
        # Year headers
        for i in range(6):  # Historical + 5 years
            col = chr(66 + i)
            if i == 0:
                ws[f'{col}3'] = "Historical"
            else:
                ws[f'{col}3'] = f"Year {i}"
            ws[f'{col}3'].font = self.header_font
            ws[f'{col}3'].fill = self.header_fill
            ws[f'{col}3'].alignment = self.center_align
        
        # Cash flow line items
        cash_flow_data = self.model_data.get("cash_flow", {})
        
        row = 4
        line_items = [
            ("Net Income", "net_income"),
            ("Depreciation & Amortization", "depreciation"),
            ("Change in Working Capital", "change_in_working_capital"),
            ("Other Operating Activities", "other_operating_activities"),
            ("Cash Flow from Operations", "operating_cash_flow"),
            ("Capital Expenditures", "capex"),
            ("Other Investing Activities", "other_investing_activities"),
            ("Cash Flow from Investing", "investing_cash_flow"),
            ("Debt Issuance (Repayment)", "debt_cash_flow"),
            ("Equity Issuance (Repurchase)", "equity_cash_flow"),
            ("Dividends Paid", "dividends"),
            ("Other Financing Activities", "other_financing_activities"),
            ("Cash Flow from Financing", "financing_cash_flow"),
            ("Net Change in Cash", "net_change_in_cash"),
            ("Free Cash Flow", "free_cash_flow")
        ]
        
        for label, key in line_items:
            ws[f'A{row}'] = label
            
            # Add data for each year
            if isinstance(cash_flow_data, dict) and key in cash_flow_data:
                for i in range(6):  # Historical + 5 years
                    col = chr(66 + i)
                    value = cash_flow_data[key].get(str(i), 0)
                    ws[f'{col}{row}'] = value
                    ws[f'{col}{row}'].number_format = '"$"#,##0.0,,"M"'
            
            # Apply styles
            if key in ["operating_cash_flow", "investing_cash_flow", "financing_cash_flow", "net_change_in_cash", "free_cash_flow"]:
                ws[f'A{row}'].font = self.subheader_font
                for i in range(6):
                    col = chr(66 + i)
                    ws[f'{col}{row}'].font = self.subheader_font
            
            row += 1
    
    def _create_valuation_sheet(self):
        """Create the valuation sheet"""
        ws = self.workbook.create_sheet("Valuation")
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Title
        ws['A1'] = f"{self.company_name} ({self.ticker}) - Valuation Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        
        # DCF Valuation
        row = 3
        ws[f'A{row}'] = "DCF Valuation"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        dcf_data = self.model_data.get("dcf_valuation", {})
        
        dcf_items = [
            ("Discount Rate (WACC)", "discount_rate", "0.0%"),
            ("Terminal Growth Rate", "terminal_growth_rate", "0.0%"),
            ("PV of Forecast Cash Flows", "pv_forecast_fcf", '"$"#,##0.0,,"M"'),
            ("Terminal Value", "terminal_value", '"$"#,##0.0,,"M"'),
            ("PV of Terminal Value", "pv_terminal_value", '"$"#,##0.0,,"M"'),
            ("Enterprise Value", "enterprise_value", '"$"#,##0.0,,"M"'),
            ("Net Debt", "net_debt", '"$"#,##0.0,,"M"'),
            ("Equity Value", "equity_value", '"$"#,##0.0,,"M"'),
            ("Shares Outstanding (millions)", "shares_outstanding", "#,##0.0"),
            ("Implied Share Price", "price_per_share", '"$"#,##0.00')
        ]
        
        for label, key, format_str in dcf_items:
            ws[f'A{row}'] = label
            
            if key in dcf_data:
                ws[f'B{row}'] = dcf_data[key]
                ws[f'B{row}'].number_format = format_str
            
            row += 1
        
        # Trading Comps Valuation
        row += 2
        ws[f'A{row}'] = "Trading Comps Valuation"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        comps_data = self.model_data.get("trading_comps_valuation", {})
        
        comps_items = [
            ("Forward EBITDA", "forward_ebitda", '"$"#,##0.0,,"M"'),
            ("EV/EBITDA Multiple", "ev_to_ebitda", "0.0x"),
            ("Enterprise Value", "enterprise_value", '"$"#,##0.0,,"M"'),
            ("Net Debt", "net_debt", '"$"#,##0.0,,"M"'),
            ("Equity Value", "equity_value", '"$"#,##0.0,,"M"'),
            ("Implied Share Price", "price_per_share", '"$"#,##0.00')
        ]
        
        for label, key, format_str in comps_items:
            ws[f'A{row}'] = label
            
            if key in comps_data:
                ws[f'B{row}'] = comps_data[key]
                ws[f'B{row}'].number_format = format_str
            
            row += 1
        
        # LBO Analysis
        row += 2
        ws[f'A{row}'] = "LBO Analysis"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        lbo_data = self.model_data.get("lbo_valuation", {})
        
        lbo_items = [
            ("Holding Period (years)", "holding_period_years", "0"),
            ("Exit EV/EBITDA Multiple", "exit_multiple", "0.0x"),
            ("Entry Enterprise Value", "entry_enterprise_value", '"$"#,##0.0,,"M"'),
            ("Initial Debt", "entry_debt", '"$"#,##0.0,,"M"'),
            ("Initial Equity", "entry_equity_value", '"$"#,##0.0,,"M"'),
            ("Exit Enterprise Value", "exit_enterprise_value", '"$"#,##0.0,,"M"'),
            ("Remaining Debt", "remaining_debt", '"$"#,##0.0,,"M"'),
            ("Exit Equity Value", "exit_equity_value", '"$"#,##0.0,,"M"'),
            ("Equity IRR", "equity_irr", "0.0%"),
            ("Cash-on-Cash Multiple", "cash_on_cash", "0.0x"),
            ("Entry Debt/EBITDA", "entry_debt_to_ebitda", "0.0x"),
            ("Exit Debt/EBITDA", "exit_debt_to_ebitda", "0.0x")
        ]
        
        for label, key, format_str in lbo_items:
            ws[f'A{row}'] = label
            
            if key in lbo_data:
                ws[f'B{row}'] = lbo_data[key]
                ws[f'B{row}'].number_format = format_str
            
            row += 1
    
    def _create_capital_structure_sheet(self):
        """Create the capital structure analysis sheet"""
        ws = self.workbook.create_sheet("Capital Structure")
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Title
        ws['A1'] = f"{self.company_name} ({self.ticker}) - Capital Structure Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:I1')
        
        # Headers
        row = 3
        headers = [
            "Debt/EBITDA", 
            "Debt/Capital", 
            "Debt ($M)", 
            "Equity Value ($M)", 
            "Enterprise Value ($M)", 
            "WACC", 
            "Credit Rating", 
            "Equity IRR", 
            "Share Price"
        ]
        
        for i, header in enumerate(headers):
            col = chr(65 + i)  # A, B, C, ...
            ws[f'{col}{row}'] = header
            ws[f'{col}{row}'].font = self.header_font
            ws[f'{col}{row}'].fill = self.header_fill
            ws[f'{col}{row}'].alignment = self.center_align
        
        # Capital structure grid data
        cap_structure_data = self.model_data.get("capital_structure_grid", [])
        
        row = 4
        for scenario in cap_structure_data:
            ws[f'A{row}'] = scenario.get("debt_to_ebitda", 0)
            ws[f'A{row}'].number_format = '0.0x'
            
            ws[f'B{row}'] = scenario.get("debt_to_capital", 0)
            ws[f'B{row}'].number_format = '0.0%'
            
            ws[f'C{row}'] = scenario.get("debt", 0) / 1_000_000  # Convert to millions
            ws[f'C{row}'].number_format = '"$"#,##0'
            
            ws[f'D{row}'] = scenario.get("equity_value", 0) / 1_000_000  # Convert to millions
            ws[f'D{row}'].number_format = '"$"#,##0'
            
            ws[f'E{row}'] = scenario.get("enterprise_value", 0) / 1_000_000  # Convert to millions
            ws[f'E{row}'].number_format = '"$"#,##0'
            
            ws[f'F{row}'] = scenario.get("wacc", 0)
            ws[f'F{row}'].number_format = '0.0%'
            
            ws[f'G{row}'] = scenario.get("credit_rating", "")
            
            ws[f'H{row}'] = scenario.get("equity_irr", 0)
            ws[f'H{row}'].number_format = '0.0%'
            
            ws[f'I{row}'] = scenario.get("share_price", 0)
            ws[f'I{row}'].number_format = '"$"#,##0.00'
            
            row += 1
        
        # Add chart
        self._add_capital_structure_chart(ws, len(cap_structure_data))
    
    def _add_capital_structure_chart(self, ws, data_rows):
        """Add capital structure chart to visualize the debt/WACC tradeoff"""
        # Create a chart for WACC vs. Debt/EBITDA
        chart = LineChart()
        chart.title = "WACC vs. Debt/EBITDA"
        chart.style = 10
        chart.y_axis.title = "WACC"
        chart.x_axis.title = "Debt/EBITDA"
        
        # Add data
        data = Reference(ws, min_col=6, min_row=3, max_row=3+data_rows, max_col=6)
        cats = Reference(ws, min_col=1, min_row=4, max_row=3+data_rows, max_col=1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Place chart
        ws.add_chart(chart, "A" + str(data_rows + 10)) 