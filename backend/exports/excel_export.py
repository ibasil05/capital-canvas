import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import io
from typing import Dict, List, Any

# Basic Styling
HEADING_FONT = Font(bold=True, size=12)
SUBHEADING_FONT = Font(bold=True, size=11)
DEFAULT_FONT = Font(size=11)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
THIN_BORDER_SIDE = Side(style='thin')
THIN_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)

async def generate_excel_export(model_results_data: Dict[str, Any]) -> bytes:
    """
    Generates an Excel file from model results data.
    FR-7: Three-statement model, Valuation views (DCF, Trading Comps, LBO)
    """
    workbook = Workbook()
    workbook.remove(workbook.active) # Remove default sheet

    # --- Summary Sheet --- 
    summary_sheet = workbook.create_sheet(title="Summary")
    _populate_summary_sheet(summary_sheet, model_results_data)

    # --- Financial Statements Sheets --- 
    # model_results_data['financial_statements'] is a list of records
    # Each record has: year, is_historical, revenue, gross_profit, ebitda, etc.
    financial_statements = model_results_data.get('financial_statements', [])
    
    if financial_statements:
        is_sheet = workbook.create_sheet(title="Income Statement")
        _populate_financial_statement_sheet(is_sheet, financial_statements, statement_type='income_statement')

        bs_sheet = workbook.create_sheet(title="Balance Sheet")
        _populate_financial_statement_sheet(bs_sheet, financial_statements, statement_type='balance_sheet')

        cf_sheet = workbook.create_sheet(title="Cash Flow Statement")
        _populate_financial_statement_sheet(cf_sheet, financial_statements, statement_type='cash_flow_statement')

    # --- Valuation Sheets (Placeholders for now) ---
    dcf_sheet = workbook.create_sheet(title="DCF Valuation")
    # _populate_dcf_sheet(dcf_sheet, model_results_data.get('valuation', {}).get('dcf_valuation', {}), financial_statements)
    dcf_sheet["A1"] = "DCF Valuation Details (Placeholder)"

    comps_sheet = workbook.create_sheet(title="Trading Comps")
    # _populate_comps_sheet(comps_sheet, model_results_data.get('valuation', {}).get('trading_comps_valuation', {}))
    comps_sheet["A1"] = "Trading Comps Details (Placeholder)"

    lbo_sheet = workbook.create_sheet(title="LBO Analysis")
    # _populate_lbo_sheet(lbo_sheet, model_results_data.get('valuation', {}).get('lbo_analysis', {}))
    lbo_sheet["A1"] = "LBO Analysis Details (Placeholder)"

    # Adjust column widths for all sheets
    for sheet in workbook.worksheets:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    return excel_file.read()

def _write_cell(sheet, row, col, value, font=None, alignment=None, border=None, number_format=None):
    cell = sheet.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    return cell

def _populate_summary_sheet(sheet, data: Dict[str, Any]):
    row_idx = 1
    sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
    _write_cell(sheet, row_idx, 1, f"Financial Model Summary: {data.get('ticker', '')} - {data.get('company_name', '')}", font=Font(bold=True, size=14), alignment=CENTER_ALIGN)
    row_idx += 2

    # Key Assumptions
    _write_cell(sheet, row_idx, 1, "Key Assumptions", font=HEADING_FONT)
    row_idx += 1
    assumptions = data.get('assumptions', {})
    key_assumptions_map = {
        "Tax Rate": assumptions.get("tax_rate"),
        "Discount Rate (WACC)": assumptions.get("discount_rate"),
        "Terminal Growth Rate": assumptions.get("terminal_growth_rate"),
        "Risk-Free Rate": assumptions.get("risk_free_rate"),
        "Market Risk Premium": assumptions.get("market_premium"),
        # Add more as needed from default_assumptions.yml or model inputs
    }
    for key, val in key_assumptions_map.items():
        _write_cell(sheet, row_idx, 1, key, font=DEFAULT_FONT, alignment=LEFT_ALIGN)
        if isinstance(val, (float)) and (key.endswith("Rate") or key.endswith("Premium") or key.endswith("WACC)")):
             _write_cell(sheet, row_idx, 2, val, font=DEFAULT_FONT, alignment=RIGHT_ALIGN, number_format='0.00%')
        else:
            _write_cell(sheet, row_idx, 2, val if val is not None else "N/A", font=DEFAULT_FONT, alignment=RIGHT_ALIGN)
        row_idx += 1
    row_idx += 1

    # Valuation Summary
    _write_cell(sheet, row_idx, 1, "Valuation Summary", font=HEADING_FONT)
    row_idx += 1
    valuation = data.get('valuation', {})
    dcf_results = valuation.get('dcf_valuation', {})
    comps_results = valuation.get('trading_comps_valuation', {})
    lbo_results = valuation.get('lbo_analysis', {})
    
    valuation_summary_map = {
        "DCF Implied Share Price": dcf_results.get("price_per_share"),
        "Trading Comps Implied Share Price": comps_results.get("price_per_share"),
        "LBO Implied Equity IRR": lbo_results.get("implied_irr") if lbo_results else None,
        "Current Market Price": data.get('company_data', {}).get('profile',{}).get('price') # Assuming it might be here
    }
    for key, val in valuation_summary_map.items():
        _write_cell(sheet, row_idx, 1, key, font=DEFAULT_FONT, alignment=LEFT_ALIGN)
        if isinstance(val, (float, int)) and "Price" in key : 
            _write_cell(sheet, row_idx, 2, val, font=DEFAULT_FONT, alignment=RIGHT_ALIGN, number_format='#,##0.00')
        elif isinstance(val, (float, int)) and "IRR" in key:
            _write_cell(sheet, row_idx, 2, val, font=DEFAULT_FONT, alignment=RIGHT_ALIGN, number_format='0.00%')
        else:
            _write_cell(sheet, row_idx, 2, val if val is not None else "N/A", font=DEFAULT_FONT, alignment=RIGHT_ALIGN)
        row_idx += 1

def _populate_financial_statement_sheet(sheet, financial_statements: List[Dict[str, Any]], statement_type: str):
    if not financial_statements:
        return

    headers = ["Metric"] + [f'{fs["year"]} ({'H' if fs["is_historical"] else 'F'})' for fs in financial_statements]
    for col_idx, header in enumerate(headers, 1):
        _write_cell(sheet, 1, col_idx, header, font=SUBHEADING_FONT, alignment=CENTER_ALIGN, border=THIN_BORDER)

    statement_items_map = {
        'income_statement': [
            ("Revenue", "revenue"), 
            ("Gross Profit", "gross_profit"), 
            ("EBITDA", "ebitda"),
            ("Depreciation & Amortization", "depreciation"), # Assuming D&A is in 'depreciation' field for now
            ("Operating Income (EBIT)", "operating_income"),
            ("Interest Expense", "interest_expense"),
            ("Income Before Tax", "income_before_tax"),
            ("Taxes", "taxes"),
            ("Net Income", "net_income")
        ],
        'balance_sheet': [
            ("Cash & Cash Equivalents", "cash_and_cash_equivalents"), # Need to ensure these fields exist
            ("Accounts Receivable", "accounts_receivable"),
            ("Inventory", "inventory"),
            ("Total Current Assets", "total_current_assets"),
            ("Property, Plant & Equipment, Net", "fixed_assets"), # fixed_assets from our model
            ("Total Assets", "total_assets"),
            ("Accounts Payable", "accounts_payable"),
            ("Short-Term Debt", "short_term_debt"),
            ("Total Current Liabilities", "total_current_liabilities"),
            ("Long-Term Debt", "long_term_debt"), # total_debt from our model might be this + short term
            ("Total Debt", "total_debt"),
            ("Total Liabilities", "total_liabilities"),
            ("Total Equity", "total_equity"),
            ("Total Liabilities & Equity", "total_liabilities_and_equity")
        ],
        'cash_flow_statement': [
            ("Net Income", "net_income"),
            ("Depreciation & Amortization", "depreciation"),
            ("Change in Working Capital", "change_in_working_capital"),
            ("Operating Cash Flow", "operating_cash_flow"),
            ("Capital Expenditures", "capex"),
            ("Investing Cash Flow", "investing_cash_flow"), # May need to derive
            ("Financing Cash Flow", "financing_cash_flow"), # May need to derive
            ("Net Change in Cash", "net_change_in_cash"), # May need to derive
            ("Free Cash Flow (FCF)", "free_cash_flow")
        ]
    }

    items_to_display = statement_items_map.get(statement_type, [])
    row_idx = 2
    for item_name, item_key in items_to_display:
        _write_cell(sheet, row_idx, 1, item_name, font=DEFAULT_FONT, alignment=LEFT_ALIGN, border=THIN_BORDER)
        for col_idx, fs_period in enumerate(financial_statements, 2):
            value = fs_period.get(item_key)
            is_percentage = item_name.endswith("Margin") or item_name.endswith("Rate")
            num_format = '0.00%' if is_percentage else '#,##0;(#,##0)' # Basic accounting format, negative in parens
            if isinstance(value, (int, float)):
                 _write_cell(sheet, row_idx, col_idx, value, font=DEFAULT_FONT, alignment=RIGHT_ALIGN, border=THIN_BORDER, number_format=num_format)
            else:
                 _write_cell(sheet, row_idx, col_idx, "N/A" if value is None else value , font=DEFAULT_FONT, alignment=RIGHT_ALIGN, border=THIN_BORDER)
        row_idx +=1

    # Add key metrics/ratios at the bottom of each statement if applicable
    # Example for Income Statement:
    if statement_type == 'income_statement':
        row_idx += 1 # Spacer row
        metrics_to_display = [
            ("Revenue Growth Rate", "growth_rate", '0.00%'),
            ("Gross Margin", "gross_margin", '0.00%'),
            ("EBITDA Margin", "ebitda_margin", '0.00%'),
            ("Net Income Margin", "net_income_margin", '0.00%') # Needs calculation if not present
        ]
        for metric_name, metric_key, num_format in metrics_to_display:
            _write_cell(sheet, row_idx, 1, metric_name, font=SUBHEADING_FONT, alignment=LEFT_ALIGN, border=THIN_BORDER)
            for col_idx, fs_period in enumerate(financial_statements, 2):
                value = fs_period.get(metric_key)
                # Calculation for net income margin if not directly available
                if metric_key == "net_income_margin" and value is None:
                    revenue = fs_period.get("revenue", 0)
                    net_income = fs_period.get("net_income", 0)
                    value = net_income / revenue if revenue else 0

                if isinstance(value, (int,float)):
                    _write_cell(sheet, row_idx, col_idx, value, font=DEFAULT_FONT, alignment=RIGHT_ALIGN, border=THIN_BORDER, number_format=num_format)
                else:
                    _write_cell(sheet, row_idx, col_idx, "N/A", font=DEFAULT_FONT, alignment=RIGHT_ALIGN, border=THIN_BORDER)
            row_idx += 1 